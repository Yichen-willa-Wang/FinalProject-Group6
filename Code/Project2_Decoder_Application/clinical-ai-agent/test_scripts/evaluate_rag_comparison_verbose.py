"""
RAG Evaluation: With RAG vs Without RAG (Verbose Mode)
Shows full answers for each question
"""
import os
os.environ["TOKENIZERS_PARALLELISM"] = "false"

import sys
import torch
import time
import pandas as pd
from datetime import datetime
from typing import Dict
from langchain_community.llms import Ollama
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import Chroma

print("="*80)
print("📊 RAG Evaluation: With RAG vs Without RAG (VERBOSE)")
print("="*80)

# ============================================================================
# 25 Test Questions
# ============================================================================

questions = [
    "1. What is refractive surgery?",
    "2. Can I have both eyes treated at the same time?",
    "3. Who can have the refractive surgery?",
    "4. What refractive defects can refractive surgery correct?",
    "5. Will the refractive surgery patient see much better than with glasses or contact lenses?",
    "6. What are the most common complications?",
    "7. Is any pain felt after laser surgery?",
    "8. Can a person have a cataract operation after laser surgery?",
    "9. If the patient has already undergone refractive surgery, is it possible to undergo surgery again in the future?",
    "10. Which technique is better: laser refractive surgery or intraocular lens implantation?",
    "11. Is it possible to undergo refractive surgery if other visual problems are present?",
    "12. What is achieved with laser surgery?",
    "13. Is refractive surgery covered by insurance?",
    "14. How safe is refractive surgery?",
    "15. What if I blink or move during the procedure?",
    "16. How soon after the surgery will I be able to see?",
    "17. How often will I see the doctor following my surgery?",
    "18. How is PRK or LASIK likely to affect my need to use glasses or contacts when I get older?",
    "19. Can I still drive at night after refractive surgery?",
    "20. Does laser eye surgery hurt?",
    "21. Is there a suitable alternative if I have thin corneas?",
    "22. What is the difference between PRK and LASIK surgery?",
    "23. Am I a suitable candidate for laser eye surgery?",
    "24. Does SMILE surgery differ from LASIK and PRK surgery?",
    "25. Why does PRK recovery take so long?"
]

# ============================================================================
# Ollama LLM Wrapper
# ============================================================================

class OllamaLLM:
    """Ollama local LLM wrapper"""
    
    def __init__(self, model_name="deepseek-r1:1.5b"):
        self.model_name = model_name
        self.api_url = "http://localhost:11434/api/generate"
    
    def generate(self, prompt: str, max_tokens=400) -> str:
        """Generate response using Ollama"""
        import requests
        
        payload = {
            "model": self.model_name,
            "prompt": prompt,
            "stream": False,
            "options": {
                "temperature": 0.7,
                "num_predict": max_tokens
            }
        }
        
        response = requests.post(self.api_url, json=payload, timeout=300)
        
        if response.status_code == 200:
            return response.json().get('response', '')
        else:
            raise Exception(f"Ollama error: {response.status_code}")

# ============================================================================
# Test Functions
# ============================================================================

def test_without_rag(llm: OllamaLLM, question: str) -> Dict:
    """Test LLM without RAG"""
    
    prompt = f"""You are a medical assistant. Answer this question about refractive surgery based on your general knowledge.

Question: {question}

Provide a clear, professional answer (150-200 words):"""
    
    start = time.time()
    try:
        answer = llm.generate(prompt, max_tokens=300)
        elapsed = time.time() - start
        
        return {
            'answer': answer,
            'time': elapsed,
            'status': 'Success',
            'sources': 'General Knowledge'
        }
    except Exception as e:
        return {
            'answer': f"Error: {str(e)}",
            'time': time.time() - start,
            'status': 'Error',
            'sources': 'N/A'
        }

def test_with_rag(llm: OllamaLLM, vectordb, question: str) -> Dict:
    """Test LLM with RAG"""
    
    start = time.time()
    
    try:
        # Retrieve documents
        docs = vectordb.similarity_search(question, k=2)
        
        if not docs:
            context = "No relevant documents found."
            sources = "None"
        else:
            context_parts = []
            sources_list = []
            
            for i, doc in enumerate(docs, 1):
                context_parts.append(f"[Reference {i}]\n{doc.page_content[:500]}")
                source = doc.metadata.get('source', 'Unknown')
                page = doc.metadata.get('page', 'N/A')
                sources_list.append(f"{source} (p.{page})")
            
            context = "\n\n".join(context_parts)
            sources = "; ".join(sources_list)
        
        # Build RAG prompt
        prompt = f"""You are a professional medical assistant. Answer based on the provided medical references.

Medical References:
{context}

Question: {question}

Instructions:
1. Answer based on the references
2. Be clear and professional
3. Keep answer to 150-200 words

Answer:"""
        
        answer = llm.generate(prompt, max_tokens=300)
        elapsed = time.time() - start
        
        return {
            'answer': answer,
            'time': elapsed,
            'status': 'Success',
            'sources': sources
        }
        
    except Exception as e:
        return {
            'answer': f"Error: {str(e)}",
            'time': time.time() - start,
            'status': 'Error',
            'sources': 'N/A'
        }

# ============================================================================
# Main Execution
# ============================================================================

print(f"\n📝 Total questions: {len(questions)}\n")

# Initialize
print("🔄 Initializing DeepSeek-R1:1.5b...")
llm = OllamaLLM(model_name="deepseek-r1:1.5b")

try:
    test = llm.generate("Hello", max_tokens=10)
    print(f"✅ LLM ready\n")
except Exception as e:
    print(f"❌ Error: {e}")
    sys.exit(1)

print("🔄 Loading embeddings on GPU...")
embeddings = HuggingFaceEmbeddings(
    model_name="BAAI/bge-base-zh-v1.5",
    model_kwargs={'device': 'cuda'}
)

print("🔄 Loading vector database...")
vectordb = Chroma(
    persist_directory="/home/ubuntu/data/vector_db",
    embedding_function=embeddings
)
print(f"✅ Vector DB: {vectordb._collection.count()} chunks\n")

# Run evaluation
print("="*80)
print("🧪 STARTING EVALUATION")
print("="*80 + "\n")

results = []

for i, question in enumerate(questions, 1):
    print("\n" + "="*80)
    print(f"QUESTION {i}/{len(questions)}")
    print("="*80)
    print(f"{question}\n")
    
    # Test WITHOUT RAG
    print("🔹 WITHOUT RAG (General Knowledge)")
    print("-"*80)
    result_no_rag = test_without_rag(llm, question)
    print(f"⏱️  Time: {result_no_rag['time']:.2f}s | Status: {result_no_rag['status']}")
    print(f"\n📝 Answer:\n{result_no_rag['answer']}\n")
    
    # Test WITH RAG
    print("🔸 WITH RAG (Retrieved Context)")
    print("-"*80)
    result_with_rag = test_with_rag(llm, vectordb, question)
    print(f"⏱️  Time: {result_with_rag['time']:.2f}s | Status: {result_with_rag['status']}")
    print(f"📚 Sources: {result_with_rag['sources']}")
    print(f"\n📝 Answer:\n{result_with_rag['answer']}\n")
    
    # Comparison
    time_diff = result_with_rag['time'] - result_no_rag['time']
    print(f"📊 Comparison:")
    print(f"   RAG overhead: {time_diff:+.2f}s")
    print(f"   Answer length: Without RAG={len(result_no_rag['answer'])} chars, With RAG={len(result_with_rag['answer'])} chars")
    
    # Store results
    results.append({
        'Question_ID': i,
        'Question': question,
        'Answer_Without_RAG': result_no_rag['answer'],
        'Time_Without_RAG': round(result_no_rag['time'], 2),
        'Status_Without_RAG': result_no_rag['status'],
        'Answer_With_RAG': result_with_rag['answer'],
        'Time_With_RAG': round(result_with_rag['time'], 2),
        'Status_With_RAG': result_with_rag['status'],
        'Sources_With_RAG': result_with_rag['sources'],
        'Time_Difference': round(time_diff, 2),
        'Length_Without_RAG': len(result_no_rag['answer']),
        'Length_With_RAG': len(result_with_rag['answer'])
    })

# Generate Excel
print("\n" + "="*80)
print("📊 GENERATING EXCEL REPORT")
print("="*80)

df = pd.DataFrame(results)
output_filename = f"rag_evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Comparison', index=False)
    
    ws = writer.sheets['Comparison']
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 100
    ws.column_dimensions['F'].width = 100
    ws.column_dimensions['I'].width = 60
    
    from openpyxl.styles import Alignment, Font, PatternFill
    for row in ws.iter_rows(min_row=2, max_row=len(df)+1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Summary
    summary_df = pd.DataFrame({
        'Metric': [
            'Total Questions',
            'Avg Time Without RAG',
            'Avg Time With RAG',
            'Avg RAG Overhead',
            'Avg Length Without RAG',
            'Avg Length With RAG'
        ],
        'Value': [
            len(results),
            f"{df['Time_Without_RAG'].mean():.2f}s",
            f"{df['Time_With_RAG'].mean():.2f}s",
            f"{df['Time_Difference'].mean():.2f}s",
            f"{df['Length_Without_RAG'].mean():.0f} chars",
            f"{df['Length_With_RAG'].mean():.0f} chars"
        ]
    })
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

print(f"✅ Saved: {output_filename}\n")

# Final summary
print("="*80)
print("📈 FINAL SUMMARY")
print("="*80)
print(f"\nTotal questions: {len(results)}")
print(f"Avg time without RAG: {df['Time_Without_RAG'].mean():.2f}s")
print(f"Avg time with RAG: {df['Time_With_RAG'].mean():.2f}s")
print(f"RAG overhead: {df['Time_Difference'].mean():.2f}s")
print(f"\n✅ EVALUATION COMPLETE!")
print(f"📄 Output: {output_filename}\n")
