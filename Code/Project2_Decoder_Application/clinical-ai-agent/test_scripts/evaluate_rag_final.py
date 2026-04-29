
"""

RAG Evaluation - Final Version

Improved prompts with consistent length and empathetic medical tone

"""

import os

os.environ["TOKENIZERS_PARALLELISM"] = "false"



import time

import pandas as pd

from datetime import datetime

from langchain_community.llms import Ollama

from langchain_community.embeddings import HuggingFaceEmbeddings

from langchain_community.vectorstores import Chroma



print("="*80)

print("📊 RAG Evaluation - 25 Questions (Enhanced Prompts)")

print("="*80)



# All 25 questions

questions = [

    "1. What is refractive surgery?",

    "2. Can I have both eyes treated at the same time?",

    "3. Who can have the refractive surgery?",

    "4. What refractive defects can refractive surgery correct?",

    "5. Will the refractive surgery patient see much better than with glasses or contact lenses?",

    "6. What are the most common complications?",

    "7. Is any pain felt after laser surgery?",

    "8. Can a person have a cataract operation after laser surgery?",

    "9. If the patient has already undergone refractive surgery, is it possible to undergo surgery again in the future?",

    "10. Which technique is better: laser refractive surgery or intraocular lens implantation?",

    "11. Is it possible to undergo refractive surgery if other visual problems are present?",

    "12. What is achieved with laser surgery?",

    "13. Is refractive surgery covered by insurance?",

    "14. How safe is refractive surgery?",

    "15. What if I blink or move during the procedure?",

    "16. How soon after the surgery will I be able to see?",

    "17. How often will I see the doctor following my surgery?",

    "18. How is PRK or LASIK likely to affect my need to use glasses or contacts when I get older?",

    "19. Can I still drive at night after refractive surgery?",

    "20. Does laser eye surgery hurt?",

    "21. Is there a suitable alternative if I have thin corneas?",

    "22. What is the difference between PRK and LASIK surgery?",

    "23. Am I a suitable candidate for laser eye surgery?",

    "24. Does SMILE surgery differ from LASIK and PRK surgery?",

    "25. Why does PRK recovery take so long?"

]



# Enhanced medical prompt template

MEDICAL_SYSTEM_PROMPT = """You are a compassionate and knowledgeable ophthalmologist. Your role is to:

- Answer patient questions with empathy and care

- Provide accurate, evidence-based medical information

- Use clear, professional language that patients can understand

- Address patient concerns and alleviate their worries

- Maintain clinical rigor while being approachable

- Always remind patients to consult their doctor for personalized advice



Keep your response between 150-200 words."""



# Initialize

print(f"\n🔄 Loading components...")

embeddings = HuggingFaceEmbeddings(

    model_name="BAAI/bge-base-zh-v1.5",

    model_kwargs={'device': 'cuda'}

)



vectordb = Chroma(

    persist_directory="/home/ubuntu/data/vector_db",

    embedding_function=embeddings

)



llm = Ollama(model="deepseek-r1:1.5b")



print(f"✅ Ready (Vector DB: {vectordb._collection.count()} chunks)\n")



# Store results

results = []



# Process each question

for i, question in enumerate(questions, 1):

    print("\n" + "="*80)

    print(f"QUESTION {i}/25")

    print("="*80)

    print(f"{question}\n")

    

    # ========================================

    # WITHOUT RAG (Consistent length prompt)

    # ========================================

    print("🔹 WITHOUT RAG (General Knowledge)")

    print("-"*80)

    

    prompt_no_rag = f"""{MEDICAL_SYSTEM_PROMPT}



Patient Question: {question}



Provide your professional medical response:"""

    

    start = time.time()

    try:

        answer_no_rag = llm.invoke(prompt_no_rag)

        time_no_rag = time.time() - start

        status_no_rag = "Success"

    except Exception as e:

        answer_no_rag = f"Error: {str(e)}"

        time_no_rag = time.time() - start

        status_no_rag = "Error"

    

    print(f"⏱️  Time: {time_no_rag:.2f}s")

    print(f"📏 Length: {len(answer_no_rag)} chars")

    print(f"\n📝 Answer:")

    print(answer_no_rag)

    print()

    

    # ========================================

    # WITH RAG (Same structure + context)

    # ========================================

    print("🔸 WITH RAG (Retrieved Medical References)")

    print("-"*80)

    

    start = time.time()

    try:

        # Retrieve documents

        docs = vectordb.similarity_search(question, k=2)

        

        if docs:

            print(f"📚 Retrieved {len(docs)} documents:")

            

            context_parts = []

            sources_list = []

            

            for j, doc in enumerate(docs, 1):

                source = doc.metadata.get('source', 'Unknown')

                page = doc.metadata.get('page', 'N/A')

                

                print(f"   {j}. {source} (Page {page})")

                

                # SHORT CONTEXT: 300 chars max to keep prompt length similar

                context_parts.append(f"[Reference {j}]\n{doc.page_content[:300]}")

                sources_list.append(f"{source} (p.{page})")

            

            context = "\n\n".join(context_parts)

            sources_str = "; ".join(sources_list)

            

            print(f"\n📏 Context length: {len(context)} chars")

            

        else:

            context = "No specific medical references found in the knowledge base for this question."

            sources_str = "None"

            print("⚠️  No documents found")

        

        # Build RAG prompt with SAME structure as no-RAG

        prompt_with_rag = f"""{MEDICAL_SYSTEM_PROMPT}



Medical References from Knowledge Base:

{context}



Patient Question: {question}



Based on the above medical references, provide your professional medical response:"""

        

        print(f"📏 Total prompt length: {len(prompt_with_rag)} chars\n")

        

        # Generate answer

        answer_with_rag = llm.invoke(prompt_with_rag)

        time_with_rag = time.time() - start

        status_with_rag = "Success"

        

    except Exception as e:

        answer_with_rag = f"Error: {str(e)}"

        time_with_rag = time.time() - start

        status_with_rag = "Error"

        sources_str = "N/A"

    

    print(f"⏱️  Time: {time_with_rag:.2f}s")

    print(f"📏 Length: {len(answer_with_rag)} chars")

    print(f"\n📝 Answer:")

    print(answer_with_rag)

    print()

    

    # Comparison

    print("📊 Comparison:")

    print(f"   Time difference: {time_with_rag - time_no_rag:+.2f}s")

    print(f"   Prompt length: No RAG={len(prompt_no_rag)} chars, With RAG={len(prompt_with_rag)} chars")

    print(f"   Answer length: No RAG={len(answer_no_rag)} chars, With RAG={len(answer_with_rag)} chars")

    

    # Store results

    results.append({

        'ID': i,

        'Question': question,

        'Answer_NoRAG': answer_no_rag,

        'Time_NoRAG': round(time_no_rag, 2),

        'Status_NoRAG': status_no_rag,

        'Length_NoRAG': len(answer_no_rag),

        'Prompt_Length_NoRAG': len(prompt_no_rag),

        'Answer_WithRAG': answer_with_rag,

        'Time_WithRAG': round(time_with_rag, 2),

        'Status_WithRAG': status_with_rag,

        'Sources': sources_str,

        'Length_WithRAG': len(answer_with_rag),

        'Prompt_Length_WithRAG': len(prompt_with_rag),

        'TimeDiff': round(time_with_rag - time_no_rag, 2),

        'PromptLengthDiff': len(prompt_with_rag) - len(prompt_no_rag)

    })



# Save to Excel

print("\n" + "="*80)

print("📊 Generating Excel Report")

print("="*80)



df = pd.DataFrame(results)

filename = f"rag_evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"



with pd.ExcelWriter(filename, engine='openpyxl') as writer:

    df.to_excel(writer, sheet_name='Results', index=False)

    

    ws = writer.sheets['Results']

    

    # Column widths

    ws.column_dimensions['A'].width = 8   # ID

    ws.column_dimensions['B'].width = 80  # Question

    ws.column_dimensions['C'].width = 100 # Answer_NoRAG

    ws.column_dimensions['D'].width = 12  # Time_NoRAG

    ws.column_dimensions['E'].width = 12  # Status_NoRAG

    ws.column_dimensions['F'].width = 12  # Length_NoRAG

    ws.column_dimensions['G'].width = 15  # Prompt_Length_NoRAG

    ws.column_dimensions['H'].width = 100 # Answer_WithRAG

    ws.column_dimensions['I'].width = 12  # Time_WithRAG

    ws.column_dimensions['J'].width = 12  # Status_WithRAG

    ws.column_dimensions['K'].width = 60  # Sources

    ws.column_dimensions['L'].width = 12  # Length_WithRAG

    ws.column_dimensions['M'].width = 15  # Prompt_Length_WithRAG

    ws.column_dimensions['N'].width = 12  # TimeDiff

    ws.column_dimensions['O'].width = 15  # PromptLengthDiff

    

    # Text wrapping

    from openpyxl.styles import Alignment, Font, PatternFill

    for row in ws.iter_rows(min_row=2, max_row=len(df)+1):

        for cell in row:

            cell.alignment = Alignment(wrap_text=True, vertical='top')

    

    # Header formatting

    for cell in ws[1]:

        cell.font = Font(bold=True)

        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")



print(f"✅ Saved: {filename}\n")



# Final summary

print("="*80)

print("📈 SUMMARY")

print("="*80)



success_no_rag = len([r for r in results if r['Status_NoRAG'] == 'Success'])

success_with_rag = len([r for r in results if r['Status_WithRAG'] == 'Success'])



print(f"\nTotal questions: {len(results)}")

print(f"Success (No RAG): {success_no_rag}/{len(results)}")

print(f"Success (With RAG): {success_with_rag}/{len(results)}")



print(f"\n⏱️  Performance:")

print(f"Avg time (No RAG): {df['Time_NoRAG'].mean():.2f}s")

print(f"Avg time (With RAG): {df['Time_WithRAG'].mean():.2f}s")

print(f"RAG overhead: {df['TimeDiff'].mean():.2f}s")



print(f"\n📏 Prompt Analysis:")

print(f"Avg prompt length (No RAG): {df['Prompt_Length_NoRAG'].mean():.0f} chars")

print(f"Avg prompt length (With RAG): {df['Prompt_Length_WithRAG'].mean():.0f} chars")

print(f"Avg prompt diff: {df['PromptLengthDiff'].mean():.0f} chars")



print(f"\n📝 Answer Analysis:")

print(f"Avg answer length (No RAG): {df['Length_NoRAG'].mean():.0f} chars")

print(f"Avg answer length (With RAG): {df['Length_WithRAG'].mean():.0f} chars")



print(f"\n📄 Output file: {filename}")

print("\n" + "="*80)

print("✅ EVALUATION COMPLETE!")

print("="*80)

